import openpyxl
import os
from openpyxl_image_loader import SheetImageLoader
from PIL import Image

# Load the Excel workbook
wb = openpyxl.load_workbook("excel_path.xlsx")
sheet = wb.active

output_folder = "output_folder_path"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

start_row = 2  
end_row = 125  
start_col = 5   
end_col = 5     

# SheetImageLoader object
image_loader = SheetImageLoader(sheet)

# Each cell ranged iteration
for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        
        cell_coordinate = sheet.cell(row=row, column=col).coordinate
        # Get image
        image = image_loader.get(cell_coordinate)

        if image:
            # Get the serial number from the first column
            serial_number = sheet.cell(row=row, column=1).value

            # Save the image with the serial number as the filename
            image_path = os.path.join(output_folder, f"{serial_number}.png")
            image.save(image_path)
            print(f"Image saved: {image_path}")
